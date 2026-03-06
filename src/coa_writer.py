"""COA 엑셀 템플릿에 데이터를 채워 넣고 저장하는 모듈."""

import copy
import subprocess
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XlImage

from .config import ANALYST_NAME_MAP, COA_CELLS
from .hplc_analyzer import PurityResult


def fill_coa(
    template_path: str,
    result: PurityResult,
    output_path: str,
    manufactured_date: str = "",
    issue_no: str = "",
    issue_date: str = "",
    appearance: str = "",
    qm_manager: str = "",
    analyst_signature_path: str = "",
    qm_signature_path: str = "",
) -> str:
    """COA 템플릿에 데이터를 채우고 저장.

    Args:
        template_path: COA 엑셀 템플릿 경로
        result: HPLC 분석 결과 (PurityResult)
        output_path: 출력 파일 경로
        manufactured_date: 제조일 (예: "2026.01.05")
        issue_no: 발행번호 (예: "COAN4010001-00")
        issue_date: 발행일 (없으면 오늘 날짜)
        appearance: 외관 (예: "Yellowish powder")
        qm_manager: QM Manager 이름
        analyst_signature_path: 분석자 서명 이미지 경로
        qm_signature_path: QM Manager 서명 이미지 경로

    Returns:
        저장된 파일 경로
    """
    wb = openpyxl.load_workbook(template_path)

    # 첫 번째 시트 사용 (템플릿 시트)
    template_sheet = wb.worksheets[0]
    ws = template_sheet

    # Item
    ws[COA_CELLS["item"]] = result.item_name

    # Lot No.
    ws[COA_CELLS["lot_no"]] = result.lot_no

    # Manufactured Date
    if manufactured_date:
        ws[COA_CELLS["manufactured_date"]] = manufactured_date

    # Issue No.
    if issue_no:
        ws[COA_CELLS["issue_no"]] = issue_no

    # Issue Date (기본값: 오늘)
    if issue_date:
        ws[COA_CELLS["issue_date"]] = issue_date
    else:
        ws[COA_CELLS["issue_date"]] = date.today().strftime("%Y.%m.%d")

    # Appearance
    if appearance:
        ws[COA_CELLS["appearance_spec"]] = appearance
        # O21 수식이 =I21 이므로 자동으로 Result에 반영됨

    # Purity: X24에 값을 넣으면 O22 수식이 자동 계산
    ws[COA_CELLS["purity_value"]] = result.purity

    # 분석자 이름
    analyst_display = ANALYST_NAME_MAP.get(
        result.analyst_name, result.analyst_name
    )
    ws[COA_CELLS["analyst_name"]] = analyst_display

    # QM Manager
    if qm_manager:
        ws[COA_CELLS["qm_manager"]] = qm_manager

    # 서명 이미지 삽입
    if analyst_signature_path and Path(analyst_signature_path).exists():
        img = XlImage(analyst_signature_path)
        img.width = 80
        img.height = 40
        ws.add_image(img, "C37")  # 분석자 서명 위치

    if qm_signature_path and Path(qm_signature_path).exists():
        img = XlImage(qm_signature_path)
        img.width = 80
        img.height = 40
        ws.add_image(img, "M37")  # QM Manager 서명 위치

    # 시트 이름 변경 (Lot 번호 기반)
    ws.title = f"COA_{result.lot_no}"

    # 불필요한 시트 제거
    for sheet_name in wb.sheetnames:
        if sheet_name != ws.title:
            del wb[sheet_name]

    # 저장
    output_path = str(output_path)
    wb.save(output_path)
    wb.close()

    return output_path


def convert_to_pdf(xlsx_path: str, output_dir: str = "") -> str:
    """LibreOffice를 사용하여 Excel → PDF 변환.

    Args:
        xlsx_path: Excel 파일 경로
        output_dir: PDF 출력 디렉토리 (빈 문자열이면 같은 디렉토리)

    Returns:
        생성된 PDF 경로
    """
    xlsx_path = Path(xlsx_path)
    if not output_dir:
        output_dir = str(xlsx_path.parent)

    import os
    env = os.environ.copy()
    env["HOME"] = "/tmp"

    result = subprocess.run(
        [
            "libreoffice", "--headless", "--calc",
            "--convert-to", "pdf",
            "--outdir", output_dir, str(xlsx_path)
        ],
        capture_output=True, text=True, timeout=60, env=env
    )

    if result.returncode != 0:
        raise RuntimeError(
            f"PDF 변환 실패: {result.stderr}"
        )

    pdf_path = Path(output_dir) / f"{xlsx_path.stem}.pdf"
    if pdf_path.exists():
        return str(pdf_path)

    raise FileNotFoundError(f"PDF 파일이 생성되지 않음: {pdf_path}")
